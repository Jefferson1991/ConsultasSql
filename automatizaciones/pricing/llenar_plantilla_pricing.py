# =============================================================================
# Llena la plantilla de facturas del modulo Pricing (RevenueOS) con las ventas
# reales de SAP B1 (HANA, EMPAQPLAST_PROD) de 2025 y 2026.
#
# Uso:
#   1. copiar .env.example a .env y poner HANA_PASSWORD
#   2. python llenar_plantilla_pricing.py
#
# Lee  : extraer-facturas-pricing.sql  (la consulta vive ahi, no aqui)
# Toma : documentacion/Plantilla/plantilla-facturas-pricing.xlsx  (no se modifica)
# Deja : documentacion/Plantilla/facturas-pricing-2025-2026.xlsx
#
# Dependencias: hdbcli, openpyxl.
#
# Nota de despliegue: se conecta directo a HANA por hdbcli. No usa el linked
# server HANAODBC porque este equipo no tiene instancia de SQL Server local.
# =============================================================================

import os
import sys
from pathlib import Path

from hdbcli import dbapi
from openpyxl import load_workbook

AQUI = Path(__file__).resolve().parent
SQL = AQUI / "extraer-facturas-pricing.sql"

PLANTILLA = Path(
    r"C:\Users\Jefferson.LTEP91\Music\RevenewOs\documentacion\Plantilla"
    r"\plantilla-facturas-pricing.xlsx"
)
SALIDA = PLANTILLA.with_name("facturas-pricing-2025-2026.xlsx")

# Corte de la clasificacion ABC por facturacion acumulada (Pareto).
ABC_CORTE_A = 0.80
ABC_CORTE_B = 0.95

FILAS_EJEMPLO = (2, 6)  # la plantilla trae 6 filas de ejemplo bajo el encabezado


def cargar_env() -> dict:
    env_path = AQUI / ".env"
    if not env_path.exists():
        sys.exit(f"Falta {env_path}. Copie .env.example a .env y ponga HANA_PASSWORD.")
    env = {}
    for linea in env_path.read_text(encoding="utf-8").splitlines():
        linea = linea.strip()
        if linea and not linea.startswith("#") and "=" in linea:
            k, v = linea.split("=", 1)
            env[k.strip()] = v.strip()
    if not env.get("HANA_PASSWORD"):
        sys.exit("HANA_PASSWORD vacio en .env")
    return env


def traer_facturas(env: dict):
    conn = dbapi.connect(
        address=env["HANA_HOST"],
        port=int(env["HANA_PORT"]),
        user=env["HANA_USER"],
        password=env["HANA_PASSWORD"],
        currentSchema=env["HANA_SCHEMA"],
        encrypt=False,
        sslValidateCertificate=False,
    )
    try:
        cur = conn.cursor()
        cur.execute(SQL.read_text(encoding="utf-8"))
        filas = cur.fetchall()
        cur.close()
    finally:
        conn.close()
    return filas


def clasificar_abc(filas, idx_cliente: int, idx_venta: int) -> dict:
    """A/B/C por facturacion neta acumulada del cliente en todo el periodo."""
    venta = {}
    for f in filas:
        venta[f[idx_cliente]] = venta.get(f[idx_cliente], 0.0) + float(f[idx_venta] or 0)

    total = sum(v for v in venta.values() if v > 0)
    if total <= 0:
        return {c: "C" for c in venta}

    abc, acumulado = {}, 0.0
    for cliente, monto in sorted(venta.items(), key=lambda kv: kv[1], reverse=True):
        acumulado += max(monto, 0.0)
        share = acumulado / total
        abc[cliente] = "A" if share <= ABC_CORTE_A else ("B" if share <= ABC_CORTE_B else "C")
    return abc


def documentar(ws, filas, abc, anomalias):
    """Deja el origen y los supuestos donde el lector los va a ver."""
    fila = ws.max_row + 2
    clientes_a = sum(1 for v in abc.values() if v == "A")
    clientes_b = sum(1 for v in abc.values() if v == "B")
    notas = [
        "ORIGEN DE LOS DATOS",
        f"Fuente: SAP B1 sobre HANA, esquema EMPAQPLAST_PROD (produccion), tablas OINV + INV1.",
        f"Periodo: facturas con DocDate entre 2025-01-01 y 2026-12-31. Extraido el 2026-08-14.",
        f"Filtro: CANCELED='N' (excluye facturas canceladas y documentos de cancelacion).",
        f"Alcance: {len(filas):,} lineas de factura. Solo facturas de venta; las notas de",
        "   credito (ORIN/RIN1) quedan fuera por decision del solicitante.",
        "",
        "COMO SE ARMO CADA COLUMNA",
        "canal = grupo de cliente en B1 (OCRG.GroupName): C. PUBLICO, MINORISTAS,",
        "   MAYORISTAS, INDUSTRIA, SEMIMAYORISTAS, CUENTA CLAVE, CONSUMO.",
        f"tipo_cliente = clasificacion ABC calculada aqui, por facturacion neta acumulada",
        f"   del cliente en el periodo (Pareto): A hasta {ABC_CORTE_A:.0%}, B hasta {ABC_CORTE_B:.0%}, C el resto.",
        f"   Resultado: {clientes_a} clientes A, {clientes_b} B, {len(abc)-clientes_a-clientes_b} C.",
        "   No es un campo de SAP: en B1 el campo U_EMPA_PRIORIDAD esta vacio.",
        "precio_lista = INV1.PriceBefDi (precio antes de descuento de linea).",
        "descuento = precio_lista - INV1.Price * (1 - OINV.DiscPrcnt/100). Incluye el",
        "   descuento de cabecera, que en B1 NO viene aplicado en la linea. Se verifico",
        "   contra produccion que SUM(LineTotal*(1-DiscPrcnt/100)) + VatSum = DocTotal.",
        "costo_unitario = INV1.StockPrice (costo de inventario al momento de facturar).",
        "moneda = INV1.Currency. Todo el periodo esta en USD.",
        "",
        "COLUMNA SIN DATO",
        "inversion_comercial queda VACIA a proposito: SAP B1 no registra inversion",
        "   comercial por linea. Se reviso U_FLETE_INTERNA, U_Estibaje y U_SEGURO_INTERNA",
        "   y estan en cero en las 14.065 facturas del periodo. No se invento un valor.",
    ]
    if anomalias:
        notas += ["", "SALVEDADES"] + anomalias
    for i, texto in enumerate(notas):
        ws.cell(row=fila + i, column=1, value=texto)


def main():
    env = cargar_env()
    print("Consultando HANA...")
    filas = traer_facturas(env)
    print(f"  {len(filas):,} lineas de factura")

    # indices segun el SELECT del .sql
    (I_NUM, I_FECHA, I_CLI, I_NOM, I_CANAL, I_SKU, I_DESC,
     I_CANT, I_LISTA, I_DTO, I_COSTO, I_MON, I_VENTA) = range(13)

    abc = clasificar_abc(filas, I_CLI, I_VENTA)

    anomalias = []
    n_dto_neg = sum(1 for f in filas if float(f[I_DTO] or 0) < -0.0001)
    n_costo_cero = sum(1 for f in filas if not f[I_COSTO] or float(f[I_COSTO]) == 0)
    n_cant_neg = sum(1 for f in filas if float(f[I_CANT] or 0) < 0)
    if n_dto_neg:
        anomalias.append(
            f"{n_dto_neg} lineas con descuento negativo: se facturo por encima del precio"
            " de lista. Se dejan tal cual, no se truncan a cero."
        )
    if n_costo_cero:
        anomalias.append(
            f"{n_costo_cero} lineas con costo_unitario en cero en B1 (StockPrice sin valor);"
            " el margen de esas lineas queda sobrestimado."
        )
    if n_cant_neg:
        anomalias.append(f"{n_cant_neg} lineas con cantidad negativa dentro de la factura.")

    print("Escribiendo Excel...")
    wb = load_workbook(PLANTILLA)
    ws = wb["facturas"]
    ws.delete_rows(FILAS_EJEMPLO[0], FILAS_EJEMPLO[1])

    fmt = "#,##0.00"
    for i, f in enumerate(filas):
        r = i + 2
        ws.cell(row=r, column=1, value=str(f[I_NUM]))
        ws.cell(row=r, column=2, value=f[I_FECHA])
        ws.cell(row=r, column=3, value=f[I_CLI])
        ws.cell(row=r, column=4, value=f[I_NOM])
        ws.cell(row=r, column=5, value=f[I_CANAL])
        ws.cell(row=r, column=6, value=abc.get(f[I_CLI], "C"))
        ws.cell(row=r, column=7, value=f[I_SKU])
        ws.cell(row=r, column=8, value=f[I_DESC])
        for col, val in ((9, f[I_CANT]), (10, f[I_LISTA]), (11, f[I_DTO]), (13, f[I_COSTO])):
            c = ws.cell(row=r, column=col, value=float(val) if val is not None else None)
            c.number_format = fmt
        c = ws.cell(row=r, column=12, value=f"=J{r}-K{r}")  # precio_facturado
        c.number_format = fmt
        ws.cell(row=r, column=14).number_format = fmt        # inversion_comercial: vacia
        ws.cell(row=r, column=15, value=f[I_MON])

    documentar(wb["instrucciones"], filas, abc, anomalias)
    wb.save(SALIDA)
    print(f"OK -> {SALIDA}")
    print(f"   filas escritas: {len(filas):,} (2..{len(filas)+1})")
    for a in anomalias:
        print("   salvedad:", a)


if __name__ == "__main__":
    main()
